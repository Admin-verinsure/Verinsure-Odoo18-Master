# -*- coding: utf-8 -*-
import io
from datetime import date

from odoo import http
from odoo.http import request, content_disposition


class CustomerStatementController(http.Controller):
    """
    HTTP controller for serving the Customer Statement Excel (.xlsx) report.

    Route: GET /customer_statement/excel?wizard_id=<int>

    Why GET (not POST)?
    -------------------
    Odoo's ir.actions.act_url opens a URL in the browser via a plain
    anchor/redirect — the browser always uses GET for these.  wizard_id
    is a single small integer so there is zero risk of a 414 error.
    The heavy data (invoice lines, formatting) is computed server-side
    by loading the wizard record; nothing large travels in the URL.
    """

    @http.route(
        '/customer_statement/excel',
        type='http',
        auth='user',
        methods=['GET'],
        csrf=False,
        save_session=False,
    )
    def download_excel(self, wizard_id=None, **kwargs):
        """
        Build and stream the Excel workbook for the given wizard record.
        """
        # ── Validate wizard_id ──────────────────────────────────────────
        try:
            wizard_id = int(wizard_id or 0)
            if not wizard_id:
                raise ValueError
        except (ValueError, TypeError):
            return request.make_response('Invalid or missing wizard_id', status=400)

        wizard = request.env['customer.statement.wizard'].browse(wizard_id)
        if not wizard.exists():
            return request.make_response('Wizard record not found', status=404)

        # ── Fetch statement data (single source of truth) ───────────────
        data = wizard._get_statement_data()

        # ── Build workbook ──────────────────────────────────────────────
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return request.make_response(
                'openpyxl not installed. Run: pip install openpyxl', status=500
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Statement"

        # ── Colour palette ──────────────────────────────────────────────
        DARK_BLUE  = "1F3864"
        MID_BLUE   = "2E75B6"
        LIGHT_BLUE = "D6E4F0"
        LIGHT_GREY = "F2F2F2"
        RED        = "C00000"
        GREEN      = "375623"

        # ── Style helpers ───────────────────────────────────────────────
        def fnt(bold=False, size=10, color="000000", italic=False):
            return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def aln(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def bdr(style="thin"):
            s = Side(style=style)
            return Border(left=s, right=s, top=s, bottom=s)

        def num(cell, fmt="#,##0.00"):
            cell.number_format = fmt

        # ── Column widths ───────────────────────────────────────────────
        for i, w in enumerate([16, 24, 16, 16, 16, 16, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Row 1 — Company name ────────────────────────────────────────
        company_info = data['company_info']
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value      = company_info['name']
        c.font       = fnt(bold=True, size=14, color="FFFFFF")
        c.fill       = fill(DARK_BLUE)
        c.alignment  = aln("center")
        ws.row_dimensions[1].height = 28

        # ── Row 2 — Company address ─────────────────────────────────────
        ws.merge_cells("A2:F2")
        c = ws["A2"]
        c.value     = company_info['address']
        c.font      = fnt(size=9, color="FFFFFF", italic=True)
        c.fill      = fill(MID_BLUE)
        c.alignment = aln("center")
        ws.row_dimensions[2].height = 16

        # ── Row 3 — spacer ──────────────────────────────────────────────
        ws.row_dimensions[3].height = 6

        # ── Rows 4-6 — Statement meta ───────────────────────────────────
        ws.merge_cells("A4:C5")
        c = ws["A4"]; c.value = "STATEMENT OF ACCOUNT"
        c.font = fnt(bold=True, size=13); c.alignment = aln("left", "center")

        def meta_label(row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.font = fnt(bold=True); c.alignment = aln("right")

        def meta_value(row, col_start, col_end, text):
            ws.merge_cells(
                start_row=row, start_column=col_start,
                end_row=row,   end_column=col_end
            )
            c = ws.cell(row=row, column=col_start, value=text)
            c.font = fnt(); c.alignment = aln("left")

        meta_label(4, 4, "Customer:")
        meta_value(4, 5, 6, data['partner']['name'])
        meta_label(5, 4, "Period:")
        meta_value(5, 5, 6, f"{data['start_date']}  →  {data['end_date']}")
        meta_label(6, 4, "As of:")
        meta_value(6, 5, 6, date.today().strftime("%d/%m/%Y"))
        for r in [4, 5, 6]:
            ws.row_dimensions[r].height = 18

        # ── Row 7 — spacer ──────────────────────────────────────────────
        ws.row_dimensions[7].height = 6

        # ── Row 8 — Opening balance ─────────────────────────────────────
        ws.merge_cells("A8:F8")
        c = ws["A8"]
        c.value     = "Opening Balance (before period)"
        c.font      = fnt(bold=True, color=DARK_BLUE)
        c.fill      = fill(LIGHT_BLUE); c.border = bdr()
        c.alignment = aln("left")

        ob = ws["G8"]
        ob.value     = data['opening_balance']
        ob.font      = fnt(bold=True, color=DARK_BLUE)
        ob.fill      = fill(LIGHT_BLUE); ob.border = bdr()
        ob.alignment = aln("right")
        num(ob)
        ws.row_dimensions[8].height = 18

        # ── Row 9 — Header ──────────────────────────────────────────────
        for col, hdr in enumerate(
            ["Date", "Document No.", "Type", "Debit", "Credit", "Paid", "Balance"], 1
        ):
            c = ws.cell(row=9, column=col, value=hdr)
            c.font = fnt(bold=True, color="FFFFFF")
            c.fill = fill(DARK_BLUE); c.border = bdr()
            c.alignment = aln("center")
        ws.row_dimensions[9].height = 20

        # ── Data rows ───────────────────────────────────────────────────
        row = 10
        for i, line in enumerate(data['lines']):
            bg = fill(LIGHT_GREY) if i % 2 == 0 else fill("FFFFFF")
            is_cn = line['move_type'] == 'out_refund'

            def cell(col, value, h="left", bold=False, color="000000"):
                c = ws.cell(row=row, column=col, value=value)
                c.font = fnt(bold=bold, color=color)
                c.fill = bg; c.border = bdr(); c.alignment = aln(h)
                return c

            cell(1, line['date'],       h="center")
            cell(2, line['name'],       h="left")
            cell(3, line['type_label'], h="center",
                 color=(RED if is_cn else GREEN))

            if line['debit']:
                c = cell(4, line['debit'], h="right"); num(c)
            else:
                cell(4, "", h="right")

            if line['credit']:
                c = cell(5, line['credit'], h="right", color=RED); num(c)
            else:
                cell(5, "", h="right")

            if line.get('paid'):
                c = cell(6, line['paid'], h="right", color=MID_BLUE); num(c)
            else:
                cell(6, "", h="right")

            bal = line['running_balance']
            c = cell(7, bal, h="right", bold=True,
                     color=(RED if bal < 0 else "000000"))
            num(c)

            ws.row_dimensions[row].height = 17
            row += 1

        # ── Closing / net balance row ────────────────────────────────────
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="NET BALANCE DUE")
        c.font = fnt(bold=True, size=11, color="FFFFFF")
        c.fill = fill(DARK_BLUE); c.border = bdr()
        c.alignment = aln("right")

        nb = data['net_balance']
        c = ws.cell(row=row, column=7, value=nb)
        c.font = fnt(bold=True, size=11,
                     color=("FF6B6B" if nb < 0 else "FFFFFF"))
        c.fill = fill(DARK_BLUE); c.border = bdr()
        c.alignment = aln("right")
        num(c)
        ws.row_dimensions[row].height = 22

        # Freeze header rows
        ws.freeze_panes = "A10"

        # ── Stream ──────────────────────────────────────────────────────
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        partner_slug = data['partner']['name'].replace(' ', '_')
        filename = f"Customer_Statement_{partner_slug}_{date.today()}.xlsx"

        return request.make_response(
            stream.read(),
            headers=[
                ('Content-Type',
                 'application/vnd.openxmlformats-officedocument'
                 '.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename)),
            ],
        )
