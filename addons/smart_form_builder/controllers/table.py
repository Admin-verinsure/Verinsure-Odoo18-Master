import json
import csv
import io
import logging
from io import BytesIO
from odoo import http
from odoo.http import request

_logger = logging.getLogger(__name__)


class SmartFormTable(http.Controller):

    def _get_form_and_check(self, form_id):
        form = request.env["smart.form"].sudo().browse(int(form_id))
        if not form.exists():
            return None
        if request.env.user._is_public():
            return None
        return form

    def _build_columns(self, form, submissions=None):
        """Build ordered column list from form field definitions.
        Also includes any extra keys found in submission data that are not
        in field_ids (e.g. fields hidden by logic that still submitted data).
        """
        seen_keys = set()
        cols = []

        # Primary: use form field definitions (preserves order + labels)
        field_map = {}
        for f in form.field_ids:
            if f.field_type == "subheading":
                continue
            key = f.name or ("field_%s" % f.id)
            seen_keys.add(key)
            col = {
                "key":    key,
                "label":  f.label or key,
                "is_key": f.is_key_field,
                "ftype":  f.field_type,
            }
            cols.append(col)
            field_map[key] = col

        # Secondary: scan actual submission data for any extra keys
        # (covers logic-hidden fields that still submitted data)
        if submissions:
            extra_keys = []
            for sub in submissions:
                try:
                    data = json.loads(sub.data_json or "{}")
                except Exception:
                    continue
                for k in data:
                    if k not in seen_keys:
                        seen_keys.add(k)
                        extra_keys.append(k)
            for key in extra_keys:
                cols.append({
                    "key":    key,
                    "label":  key.replace("_", " ").title(),
                    "is_key": False,
                    "ftype":  "text",
                })
        return cols

    def _fmt_date(self, dt, fmt="%d/%m/%Y %H:%M:%S"):
        """Safely format an Odoo datetime field — handles False, None, and datetime."""
        if not dt:
            return ""
        try:
            return dt.strftime(fmt)
        except Exception:
            return str(dt)

    def _cell_value(self, raw, ftype, submission_id=None, escape=False):
        """Convert a raw JSON value to display string or HTML.
        If escape=True, returns plain text (for CSV/XLSX).
        If submission_id is given, file fields get a clickable download link.
        """
        if raw is None or raw == "":
            return ""
        if isinstance(raw, list):
            return ", ".join(str(v) for v in raw if v != "")
        if isinstance(raw, dict):
            return str(raw.get("label") or raw.get("value") or "")
        val = str(raw)
        if not escape and ftype == "file" and val and submission_id:
            import html as _html
            safe = _html.escape(val)
            url = "/smart_form/file/%d/%s" % (submission_id, safe)
            return '<a href="%s" target="_blank" style="color:#4f46e5;">&#128206; %s</a>' % (url, safe)
        return val

    @http.route("/smart_form/table/<int:form_id>", type="http", auth="user", website=False)
    def submission_table(self, form_id, **kw):
        form = self._get_form_and_check(form_id)
        if not form:
            return request.not_found()

        submissions = request.env["smart.form.submission"].sudo().search(
            [("form_id", "=", form.id)], order="create_date desc"
        )
        cols = self._build_columns(form, submissions)

        def esc(s):
            return (str(s)
                    .replace("&", "&amp;").replace("<", "&lt;")
                    .replace(">", "&gt;").replace('"', "&quot;"))

        # Build header cells
        th_date = '''<th class="col-date">
            <div class="th-inner">
                <span class="th-icon">&#128197;</span>
                <span class="th-label">Submitted On</span>
            </div>
        </th>'''

        th_chain = '''<th class="col-field" title="Submission Chain" style="min-width:60px;width:60px;">
            <div class="th-inner" style="min-width:auto;">
                <span class="th-icon">&#128279;</span>
                <span class="th-label">Chain</span>
            </div>
        </th>'''

        th_cols = ""
        for c in cols:
            icon = {
                "email": "&#9993;", "phone": "&#128222;", "number": "&#35;",
                "select": "&#9660;", "radio": "&#9673;", "checkbox": "&#9745;",
                "file": "&#128206;", "textarea": "&#128195;",
            }.get(c["ftype"], "&#9632;")
            key_pip = '<span class="key-pip" title="Key field">K</span>' if c["is_key"] else ""
            extra_class = " col-key" if c["is_key"] else ""
            th_cols += '''<th class="col-field%s" title="%s">
                <div class="th-inner">
                    <span class="th-icon">%s</span>
                    <span class="th-label">%s</span>
                    %s
                </div>
            </th>''' % (extra_class, esc(c["label"]), icon, esc(c["label"]), key_pip)

        # Build data rows
        rows_html = ""
        for i, sub in enumerate(submissions):
            try:
                data = json.loads(sub.data_json or "{}")
            except Exception:
                data = {}

            dt = self._fmt_date(sub.create_date, "%d %b %Y, %H:%M")
            row_class = "row-even" if i % 2 == 0 else "row-odd"
            row = '<tr class="%s">' % row_class
            row += '<td class="cell-date">%s</td>' % esc(dt)
            # Chain link cell — only show if a follow-up submission exists
            has_chain = bool(sub.session_token and sub.child_submission_ids)
            if has_chain:
                chain_url = "/smart_form/chain/%d" % sub.id
                row += ('<td style="text-align:center;padding:8px;vertical-align:middle;">'
                        '<a href="%s" title="View chain" '
                        'style="color:#4f46e5;font-size:1.1rem;text-decoration:none;">&#128279;</a>'
                        '</td>') % chain_url
            else:
                row += '<td style="text-align:center;padding:8px;"></td>'
            for c in cols:
                raw = data.get(c["key"], "")
                val = self._cell_value(raw, c["ftype"], submission_id=sub.id)
                empty_class = " cell-empty" if not val else ""
                key_class = " cell-key" if c["is_key"] else ""
                # val may contain HTML (file link) — don't double-escape it
                display_html = val if val else "<span style='color:#d1d5db;'>—</span>"
                raw_text = self._cell_value(raw, c["ftype"], escape=True)
                row += '<td class="cell-field%s%s" title="%s">%s</td>' % (
                    key_class, empty_class, esc(raw_text), display_html)
            row += "</tr>"
            rows_html += row

        if not submissions:
            span = len(cols) + 2  # +1 date +1 chain
            rows_html = '<tr><td colspan="%d" class="cell-empty-state">No submissions yet.</td></tr>' % span

        total = len(submissions)
        key_col = next((c for c in cols if c["is_key"]), None)
        export_url = "/smart_form/table/%d/export.xlsx" % form.id

        html = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>%(form_name)s — Submissions</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    body {
      font-family: 'Segoe UI', system-ui, -apple-system, Roboto, Arial, sans-serif;
      background: #f0f2f8;
      color: #1a202c;
      min-height: 100vh;
      overflow-x: auto;
    }

    /* ── TOP BAR ── */
    .topbar {
      background: linear-gradient(135deg, #4f46e5 0%%, #7c3aed 100%%);
      padding: 0 32px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      height: 64px;
      box-shadow: 0 2px 12px rgba(79,70,229,0.35);
      position: sticky;
      top: 0;
      z-index: 100;
    }
    .topbar-left { display: flex; align-items: center; gap: 14px; }
    .topbar-icon {
      width: 38px; height: 38px;
      background: rgba(255,255,255,0.18);
      border-radius: 10px;
      display: flex; align-items: center; justify-content: center;
      font-size: 1.2rem;
    }
    .topbar-title { color: #fff; font-size: 1.05rem; font-weight: 700; }
    .topbar-sub { color: rgba(255,255,255,0.65); font-size: 0.78rem; margin-top: 1px; }
    .topbar-right { display: flex; gap: 10px; align-items: center; }

    .btn {
      display: inline-flex; align-items: center; gap: 7px;
      padding: 8px 16px; border-radius: 8px;
      font-size: 0.82rem; font-weight: 600;
      cursor: pointer; text-decoration: none; border: none;
      transition: all 0.15s ease;
      white-space: nowrap;
    }
    .btn-export {
      background: #fff; color: #4f46e5;
      box-shadow: 0 1px 4px rgba(0,0,0,0.12);
    }
    .btn-export:hover { background: #f0f0ff; transform: translateY(-1px); }
    .btn-back {
      background: rgba(255,255,255,0.15); color: #fff;
      border: 1px solid rgba(255,255,255,0.3);
    }
    .btn-back:hover { background: rgba(255,255,255,0.25); }

    /* ── STATS BAR ── */
    .statsbar {
      padding: 16px 32px;
      display: flex;
      align-items: center;
      gap: 16px;
      flex-wrap: wrap;
    }
    .stat-card {
      background: #fff;
      border-radius: 10px;
      padding: 10px 18px;
      display: flex; align-items: center; gap: 10px;
      box-shadow: 0 1px 4px rgba(0,0,0,0.07);
      font-size: 0.85rem;
    }
    .stat-icon { font-size: 1.1rem; }
    .stat-label { color: #6b7280; }
    .stat-value { font-weight: 700; color: #1a202c; margin-left: 4px; }
    .stat-key-badge {
      background: linear-gradient(135deg, #4f46e5, #7c3aed);
      color: #fff;
      border-radius: 6px; padding: 3px 10px;
      font-size: 0.75rem; font-weight: 700;
      letter-spacing: 0.03em;
    }

    /* ── TABLE WRAPPER ── */
    .table-wrap {
      padding: 0 32px 32px;
      overflow-x: scroll;
      overflow-y: visible;
      -webkit-overflow-scrolling: touch;
    }
    .table-scroll-inner {
      display: inline-block;
      min-width: 100%%;
      border-radius: 14px;
      box-shadow: 0 2px 16px rgba(0,0,0,0.08);
      overflow: visible;
    }
    .table-card {
      background: #fff;
      border-radius: 14px;
      overflow: visible;
    }

    /* ── TABLE ── */
    table {
      border-collapse: collapse;
      font-size: 0.875rem;
      table-layout: auto;
      white-space: nowrap;
    }

    /* Header */
    thead {
      background: linear-gradient(135deg, #4f46e5 0%%, #6d28d9 100%%);
    }
    thead th {
      padding: 0;
      vertical-align: top;
      border-right: 1px solid rgba(255,255,255,0.12);
      position: relative;
      min-width: 140px;
      width: 160px;
      max-width: 200px;
    }
    thead th:last-child { border-right: none; }

    .th-inner {
      display: flex;
      flex-direction: column;
      align-items: flex-start;
      gap: 4px;
      padding: 12px 12px 10px;
      width: 100%%;
      height: 72px;
      overflow: hidden;
    }
    .th-icon {
      font-size: 0.95rem;
      opacity: 0.75;
      line-height: 1;
    }
    .th-label {
      color: #fff;
      font-size: 0.72rem;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      line-height: 1.35;
      white-space: normal;
      word-break: break-word;
      overflow: hidden;
      display: -webkit-box;
      -webkit-line-clamp: 3;
      -webkit-box-orient: vertical;
    }
    .key-pip {
      background: #fbbf24;
      color: #1a202c;
      border-radius: 4px;
      padding: 1px 6px;
      font-size: 0.65rem;
      font-weight: 800;
      letter-spacing: 0.06em;
    }

    /* Key column header */
    th.col-key { background: rgba(251,191,36,0.18); }
    th.col-date .th-inner { min-width: 130px; }

    /* Body rows */
    tbody tr { transition: background 0.1s; }
    tbody tr.row-even { background: #fff; }
    tbody tr.row-odd  { background: #f8f9ff; }
    tbody tr:hover    { background: #eef0ff !important; }
    tbody tr:last-child td { border-bottom: none; }

    tbody td {
      padding: 11px 14px;
      border-bottom: 1px solid #e8eaf0;
      border-right: 1px solid #f0f1f7;
      color: #374151;
      vertical-align: middle;
      max-width: 220px;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }
    tbody td:last-child { border-right: none; }

    td.cell-date {
      color: #6b7280;
      font-size: 0.82rem;
      white-space: nowrap;
      font-variant-numeric: tabular-nums;
    }
    td.cell-key {
      font-weight: 600;
      color: #4f46e5;
    }
    td.cell-empty {
      color: #d1d5db;
      font-style: italic;
    }
    td.cell-empty-state {
      text-align: center;
      padding: 48px 16px;
      color: #9ca3af;
      font-size: 0.95rem;
    }

    /* ── RESPONSIVE ── */
    @media (max-width: 768px) {
      .topbar { padding: 0 16px; }
      .statsbar, .table-wrap { padding-left: 16px; padding-right: 16px; }
      .topbar-sub { display: none; }
    }
  </style>
</head>
<body>

<!-- TOP BAR -->
<div class="topbar">
  <div class="topbar-left">
    <div class="topbar-icon">&#128203;</div>
    <div>
      <div class="topbar-title">%(form_name)s</div>
      <div class="topbar-sub">Submissions Table</div>
    </div>
  </div>
  <div class="topbar-right">
    <a href="%(export_url)s" class="btn btn-export">
      <span>&#11123;</span> Export Excel
    </a>
    <a href="javascript:history.back()" class="btn btn-back">
      <span>&#8592;</span> Back
    </a>
  </div>
</div>

<!-- STATS BAR -->
<div class="statsbar">
  <div class="stat-card">
    <span class="stat-icon">&#128203;</span>
    <span class="stat-label">Total Submissions</span>
    <span class="stat-value">%(total)s</span>
  </div>
  <div class="stat-card">
    <span class="stat-icon">&#9635;</span>
    <span class="stat-label">Fields</span>
    <span class="stat-value">%(col_count)s</span>
  </div>
  %(key_stat)s
</div>

<!-- TABLE -->
<div class="table-wrap">
  <div class="table-scroll-inner">
  <div class="table-card">
    <table>
      <thead>
        <tr>
          %(th_date)s
          %(th_chain)s
          %(th_cols)s
        </tr>
      </thead>
      <tbody>
        %(rows_html)s
      </tbody>
    </table>
  </div>
  </div>
</div>

</body>
</html>""" % {
            "form_name": esc(form.name),
            "export_url": export_url,
            "total": total,
            "col_count": len(cols),
            "key_stat": (
                '<div class="stat-card"><span class="stat-icon">&#128273;</span>'
                '<span class="stat-label">Key Field</span>'
                '<span class="stat-key-badge">%s</span></div>' % esc(key_col["label"])
            ) if key_col else "",
            "th_date": th_date,
            "th_chain": th_chain,
            "th_cols": th_cols,
            "rows_html": rows_html,
        }

        return request.make_response(
            html, [("Content-Type", "text/html; charset=utf-8")]
        )

    # ----------------------------------------------------------
    # XLSX export  (opens perfectly in Excel — no ### date issue)
    # ----------------------------------------------------------

    # ----------------------------------------------------------
    # Submission Chain view
    # ----------------------------------------------------------
    @http.route("/smart_form/chain/<int:submission_id>", type="http", auth="user", website=False)
    def submission_chain(self, submission_id, **kw):
        """Show all linked submissions in a branching chain."""
        sub = request.env["smart.form.submission"].sudo().browse(submission_id)
        if not sub.exists():
            return request.not_found()

        chain = sub.get_chain()

        def esc(s):
            return (str(s).replace("&","&amp;").replace("<","&lt;")
                          .replace(">","&gt;").replace('"',"&quot;"))

        def fmt_val(val, ftype, sub_id):
            if val is None or val == "":
                return "<span style='color:#ccc;'>&#8212;</span>"
            if isinstance(val, list):
                return ", ".join(esc(str(v)) for v in val if v)
            if isinstance(val, dict):
                return esc(str(val.get("label") or val.get("value") or ""))
            s = str(val)
            if ftype == "file":
                url = "/smart_form/file/%d/%s" % (sub_id, esc(s))
                return "<a href='%s' target='_blank' style='color:#4f46e5;'>&#128206; %s</a>" % (url, esc(s))
            if ftype == "email":
                return "<a href='mailto:%s' style='color:#4f46e5;'>%s</a>" % (esc(s), esc(s))
            if ftype == "phone":
                return "<a href='tel:%s' style='color:#4f46e5;'>%s</a>" % (esc(s), esc(s))
            return esc(s)

        STEP_COLORS = ["#4f46e5", "#7c3aed", "#0891b2", "#059669", "#d97706"]
        cards_html = ""

        for idx, s in enumerate(chain):
            try:
                data = json.loads(s.data_json or "{}")
            except Exception:
                data = {}

            label_map, type_map = {}, {}
            if s.form_id:
                for f in s.form_id.field_ids:
                    k = f.name or ("field_%s" % f.id)
                    label_map[k] = f.label or k
                    type_map[k]  = f.field_type

            known = ([f.name or ("field_%s" % f.id)
                      for f in s.form_id.field_ids if f.field_type != "subheading"]
                     if s.form_id else [])
            all_keys = known + [k for k in data if k not in known]

            rows = ""
            for i, key in enumerate(all_keys):
                if key not in data:
                    continue
                bg    = "#ffffff" if i % 2 == 0 else "#f9fafb"
                label = label_map.get(key, key.replace("_", " ").title())
                val   = fmt_val(data[key], type_map.get(key, "text"), s.id)
                rows += (
                    "<tr style='background:" + bg + ";border-bottom:1px solid #eef0f3;'>"
                    "<td style='padding:9px 14px;font-weight:600;color:#374151;font-size:0.85rem;"
                    "width:35%;white-space:nowrap;vertical-align:top;'>" + esc(label) + "</td>"
                    "<td style='padding:9px 14px;color:#1a202c;font-size:0.875rem;vertical-align:top;'>" + val + "</td>"
                    "</tr>"
                )

            color     = STEP_COLORS[idx % len(STEP_COLORS)]
            dt        = s.create_date.strftime("%d %b %Y, %H:%M") if s.create_date else ""
            form_name = esc(s.form_id.name if s.form_id else "Unknown Form")
            key_badge = ""
            if s.key_value:
                key_badge = (
                    "<span style='background:" + color + ";color:#fff;border-radius:5px;"
                    "padding:1px 8px;font-size:0.72rem;font-weight:700;margin-left:6px;'>"
                    + esc(s.key_value) + "</span>"
                )
            no_data_row = (
                "<tr><td colspan='2' style='padding:16px;color:#aaa;text-align:center;'>No data</td></tr>"
            )

            # Arrow connector between cards
            if idx > 0:
                cards_html += (
                    "<div style='text-align:center;margin:4px 0 12px;font-size:1.4rem;color:#9ca3af;'>&#8595;</div>"
                )

            cards_html += (
                "<div style='margin-bottom:20px;'>"
                  "<div style='display:flex;align-items:center;gap:12px;margin-bottom:10px;'>"
                    "<div style='background:" + color + ";color:#fff;border-radius:50%;width:32px;height:32px;"
                    "display:flex;align-items:center;justify-content:center;font-weight:700;"
                    "font-size:0.9rem;flex-shrink:0;'>" + str(idx + 1) + "</div>"
                    "<div>"
                      "<div style='font-weight:700;font-size:1rem;color:#1a202c;'>" + form_name + "</div>"
                      "<div style='font-size:0.8rem;color:#6b7280;'>" + esc(dt) + key_badge + "</div>"
                    "</div>"
                  "</div>"
                  "<div style='border-radius:10px;overflow:hidden;border:1px solid #e5e7eb;"
                  "box-shadow:0 1px 4px rgba(0,0,0,0.06);'>"
                    "<table style='width:100%;border-collapse:collapse;'>"
                      "<tbody>" + (rows or no_data_row) + "</tbody>"
                    "</table>"
                  "</div>"
                "</div>"
            )

        total  = len(chain)
        plural = "s" if total != 1 else ""

        html = (
            "<!DOCTYPE html>"
            "<html lang='en'><head>"
            "<meta charset='UTF-8'/>"
            "<meta name='viewport' content='width=device-width,initial-scale=1.0'/>"
            "<title>Submission Chain</title>"
            "<style>"
            "*{box-sizing:border-box;margin:0;padding:0;}"
            "body{font-family:'Segoe UI',Roboto,Arial,sans-serif;background:#f0f2f8;color:#1a202c;min-height:100vh;}"
            ".topbar{background:linear-gradient(135deg,#4f46e5,#7c3aed);padding:0 28px;height:60px;"
            "display:flex;align-items:center;justify-content:space-between;"
            "box-shadow:0 2px 10px rgba(79,70,229,0.3);position:sticky;top:0;z-index:100;}"
            ".topbar-title{color:#fff;font-size:1rem;font-weight:700;}"
            ".topbar-sub{color:rgba(255,255,255,0.65);font-size:0.78rem;margin-top:2px;}"
            ".btn{display:inline-flex;align-items:center;gap:6px;padding:7px 16px;border-radius:7px;"
            "font-size:0.82rem;font-weight:600;cursor:pointer;text-decoration:none;border:none;}"
            ".btn-back{background:rgba(255,255,255,0.15);color:#fff;border:1px solid rgba(255,255,255,0.3);}"
            ".body-wrap{max-width:760px;margin:0 auto;padding:28px 20px 48px;}"
            ".stat{background:#fff;border-radius:10px;padding:10px 18px;display:inline-flex;"
            "align-items:center;gap:8px;box-shadow:0 1px 4px rgba(0,0,0,0.07);"
            "font-size:0.85rem;margin-bottom:24px;}"
            "</style></head><body>"
            "<div class='topbar'>"
              "<div>"
                "<div class='topbar-title'>&#128279; Submission Chain</div>"
                "<div class='topbar-sub'>" + str(total) + " form" + plural + " completed</div>"
              "</div>"
              "<a href='javascript:history.back()' class='btn btn-back'>&#8592; Back</a>"
            "</div>"
            "<div class='body-wrap'>"
              "<div class='stat'>&#128203; <strong>" + str(total) + "</strong>"
              " <span style='color:#6b7280;'>form" + plural + " in this chain</span></div>"
              + cards_html +
            "</div>"
            "</body></html>"
        )

        return request.make_response(html, [("Content-Type", "text/html; charset=utf-8")])

    # ----------------------------------------------------------
    # File attachment download
    # ----------------------------------------------------------
    @http.route("/smart_form/file/<int:submission_id>/<string:filename>",
                type="http", auth="user", website=False)
    def download_file(self, submission_id, filename, **kw):
        """Serve a file uploaded via a form field."""
        attachment = request.env["ir.attachment"].sudo().search([
            ("res_model", "=", "smart.form.submission"),
            ("res_id",    "=", submission_id),
            ("name",      "=", filename),
        ], limit=1)
        if not attachment:
            return request.not_found()
        import base64 as _b64
        file_data = _b64.b64decode(attachment.datas or "")
        mimetype = attachment.mimetype or "application/octet-stream"
        return request.make_response(
            file_data,
            [
                ("Content-Type", mimetype),
                ("Content-Disposition", 'attachment; filename="%s"' % filename),
                ("Content-Length", str(len(file_data))),
            ],
        )

    @http.route("/smart_form/table/<int:form_id>/export.csv", type="http", auth="user", website=False)
    def export_csv(self, form_id, **kw):
        """Keep the /export.csv URL for backwards compat but now serves XLSX."""
        return self._do_export(form_id, fmt="xlsx")

    @http.route("/smart_form/table/<int:form_id>/export.xlsx", type="http", auth="user", website=False)
    def export_xlsx(self, form_id, **kw):
        return self._do_export(form_id, fmt="xlsx")

    def _do_export(self, form_id, fmt="xlsx"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side, numbers
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            return self._do_export_csv_fallback(form_id)

        form = self._get_form_and_check(form_id)
        if not form:
            return request.not_found()

        submissions = request.env["smart.form.submission"].sudo().search(
            [("form_id", "=", form.id)], order="create_date desc"
        )
        cols = self._build_columns(form, submissions)

        # Also collect columns from linked chain forms
        chain_forms = {}  # form_id -> form record, ordered by chain_depth
        for sub in submissions:
            if sub.session_token:
                chain_subs = request.env["smart.form.submission"].sudo().search(
                    [("session_token", "=", sub.session_token),
                     ("form_id", "!=", form.id)],
                    order="chain_depth asc"
                )
                for cs in chain_subs:
                    if cs.form_id.id not in chain_forms:
                        chain_forms[cs.form_id.id] = cs.form_id

        chain_cols = []  # list of (form_label_prefix, col_dict)
        for fid, cf in chain_forms.items():
            cf_subs = request.env["smart.form.submission"].sudo().search(
                [("form_id", "=", fid)], limit=50)
            for col in self._build_columns(cf, cf_subs):
                chain_cols.append((cf.name, col))

        wb = Workbook()
        ws = wb.active
        ws.title = (form.name or "Submissions")[:31]   # Excel sheet name limit

        # ── Styles ──────────────────────────────────────────────
        header_fill   = PatternFill("solid", fgColor="4F46E5")   # indigo
        key_fill      = PatternFill("solid", fgColor="7C3AED")   # purple for key col
        header_font   = Font(bold=True, color="FFFFFF", size=10)
        key_font      = Font(bold=True, color="FFFFFF", size=10)
        date_font     = Font(color="6B7280", size=9)
        key_cell_font = Font(bold=True, color="4F46E5", size=10)
        normal_font   = Font(size=10)
        center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align    = Alignment(horizontal="left",   vertical="center")
        thin_side     = Side(style="thin", color="E5E7EB")
        thin_border   = Border(left=thin_side, right=thin_side,
                               top=thin_side, bottom=thin_side)
        even_fill     = PatternFill("solid", fgColor="F8F9FF")
        odd_fill      = PatternFill("solid", fgColor="FFFFFF")

        # ── Header row ──────────────────────────────────────────
        headers = (["Submitted On"] + 
                   [c["label"] for c in cols] +
                   [("[%s] %s" % (fp, c["label"])) for fp, c in chain_cols])
        ws.row_dimensions[1].height = 36

        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            # col_idx 1=date, 2..len(cols)+1=main cols, rest=chain cols
            main_col_offset = col_idx - 2
            is_key = (0 <= main_col_offset < len(cols) and cols[main_col_offset]["is_key"])
            is_chain = col_idx > len(cols) + 1
            chain_fill = PatternFill("solid", fgColor="0891B2")  # teal for chain cols
            cell.fill = key_fill if is_key else (chain_fill if is_chain else header_fill)
            cell.font        = key_font if is_key else header_font
            cell.alignment   = center_align
            cell.border      = thin_border

        # ── Data rows ───────────────────────────────────────────
        for row_idx, sub in enumerate(submissions, start=2):
            try:
                data = json.loads(sub.data_json or "{}")
            except Exception:
                data = {}

            row_fill = even_fill if row_idx % 2 == 0 else odd_fill
            ws.row_dimensions[row_idx].height = 20

            # Date column
            dt_val = sub.create_date if sub.create_date else None
            date_cell = ws.cell(row=row_idx, column=1, value=dt_val)
            if dt_val:
                date_cell.number_format = "DD/MM/YYYY HH:MM:SS"
            date_cell.font      = date_font
            date_cell.fill      = row_fill
            date_cell.alignment = left_align
            date_cell.border    = thin_border

            # Field columns
            for col_idx, c in enumerate(cols, start=2):
                raw = data.get(c["key"], "")
                val = self._cell_value(raw, c["ftype"], escape=True)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill      = row_fill
                cell.font      = key_cell_font if c["is_key"] else normal_font
                cell.alignment = left_align
                cell.border    = thin_border

            # Chain columns — fetch linked submission data
            if chain_cols:
                chain_data_map = {}
                if sub.session_token:
                    chain_subs = request.env["smart.form.submission"].sudo().search(
                        [("session_token", "=", sub.session_token),
                         ("form_id", "!=", form.id)],
                        order="chain_depth asc"
                    )
                    for cs in chain_subs:
                        try:
                            chain_data_map[cs.form_id.id] = json.loads(cs.data_json or "{}")
                        except Exception:
                            pass
                chain_start_col = len(cols) + 2
                for cc_idx, (form_prefix, c) in enumerate(chain_cols):
                    fid = next((fid for fid, cf in chain_forms.items()
                               if cf.name == form_prefix), None)
                    cd = chain_data_map.get(fid, {}) if fid else {}
                    raw = cd.get(c["key"], "")
                    val = self._cell_value(raw, c["ftype"], escape=True)
                    cc_cell = ws.cell(row=row_idx, column=chain_start_col + cc_idx, value=val)
                    cc_cell.fill      = row_fill
                    cc_cell.font      = normal_font
                    cc_cell.alignment = left_align
                    cc_cell.border    = thin_border

        # ── Column widths ────────────────────────────────────────
        ws.column_dimensions[get_column_letter(1)].width = 20   # date

        for col_idx, c in enumerate(cols, start=2):
            label_len = len(c["label"])
            width = max(12, min(35, label_len * 1.1 + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        chain_start_col = len(cols) + 2
        for cc_idx, (form_prefix, c) in enumerate(chain_cols):
            label_len = len(c["label"]) + len(form_prefix) + 3
            width = max(14, min(40, label_len * 1.0 + 4))
            ws.column_dimensions[get_column_letter(chain_start_col + cc_idx)].width = width

        # ── Freeze top row + auto-filter ─────────────────────────
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ── Write to bytes ───────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.read()

        filename = "submissions_%s.xlsx" % (form.name or str(form.id)).replace(" ", "_")
        return request.make_response(
            xlsx_bytes,
            [
                ("Content-Type",
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", 'attachment; filename="%s"' % filename),
            ],
        )

    def _do_export_csv_fallback(self, form_id):
        """Plain CSV fallback if openpyxl is unavailable."""
        form = self._get_form_and_check(form_id)
        if not form:
            return request.not_found()
        cols = self._build_columns(form)
        submissions = request.env["smart.form.submission"].sudo().search(
            [("form_id", "=", form.id)], order="create_date desc"
        )
        output = io.StringIO()
        import csv as _csv
        writer = _csv.writer(output)
        writer.writerow(
            ["Submitted On"] +
            [c["label"] for c in cols] +
            [("[%s] %s" % (form_prefix, c["label"])) for form_prefix, c in chain_cols]
        )
        for sub in submissions:
            try:
                data = json.loads(sub.data_json or "{}")
            except Exception:
                data = {}
            writer.writerow(
                [self._fmt_date(sub.create_date)] +
                [self._cell_value(data.get(c["key"], ""), c["ftype"]) for c in cols]
            )
        filename = "submissions_%s.csv" % (form.name or str(form.id)).replace(" ", "_")
        return request.make_response(
            output.getvalue(),
            [("Content-Type", "text/csv; charset=utf-8"),
             ("Content-Disposition", 'attachment; filename="%s"' % filename)],
        )
